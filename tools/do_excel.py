from openpyxl import load_workbook
from tools import project_path
from tools.read_config import ReadConfig
# from tools.get_excel import *
# from tools.do_mysql import Do_Mysql
# query_sql = 'select max(MobilePhone) from member where MobilePhone like"156%"'
# tel_2=Do_Mysql().do_mysql(query_sql)[0]
# print(tel_2)
class Do_excle:
    # 获取表格信息


    # 读取excel中信息
    @staticmethod
    def get_information(file_name):
        wb=load_workbook(file_name)
        mode=eval(ReadConfig.get_config(project_path.test_confige_path,'MODE','mode'))
        test_data = []
        tel = wb['init'].cell(3, 2).value
        # print(tel)
        qq_1 = wb['init'].cell(2, 2).value
        # print(qq_1)
        time_side = wb['init'].cell(1, 2).value
        # print(time_side)
        for key in mode:
            sheet=wb[key]
            if mode[key] == 'all':
                for i in range(1, sheet.max_row):
                    sub_data = {}
                    sub_data['case_id'] = sheet.cell(i+1,1).value
                    sub_data['url'] = sheet.cell(i + 1, 2).value
                    sub_data['method'] = sheet.cell(i + 1, 3).value
                    if sheet.cell(i+1,4).value.find('${tel_1}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${tel_1}',str(tel))
                    elif sheet.cell(i+1,4).value.find('${time_side}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${time_side}',str(time_side))
                    elif sheet.cell(i+1,4).value.find('${qa}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${qa}',str(qq_1))
                    else:
                        sub_data['data_2'] = sheet.cell(i + 1, 4).value
                    sub_data['res'] = sheet.cell(i + 1, 6).value
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
            else:
                for i in mode[key]:
                    sub_data = {}
                    sub_data['case_id']=sheet.cell(i+1,1).value
                    sub_data['url'] = sheet.cell(i + 1, 2).value
                    sub_data['method'] = sheet.cell(i + 1, 3).value
                    if sheet.cell(i+1,4).value.find('${tel_1}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${tel_1}',str(tel))
                    elif sheet.cell(i+1,4).value.find('${time_side}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${time_side}',str(time_side))
                    elif sheet.cell(i+1,4).value.find('${qa}') != -1:
                        sub_data['data_2'] = sheet.cell(i+1,4).value.replace('${qa}',str(qq_1))
                    else:
                        sub_data['data_2'] = sheet.cell(i + 1, 4).value
                    sub_data['res'] = sheet.cell(i + 1, 6).value
                    sub_data['sheet_name'] = key
                    test_data.append(sub_data)
            sheet_2=wb['init']
            sheet_2.cell(3, 2).value=tel+1
            sheet_2.cell(2, 2).value=qq_1+1
            sheet_2.cell(1,2).value=time_side+1
            wb.save(project_path.test_path)

        return test_data

    # 写回信息进入Excel
    @staticmethod
    def write_information(file_name, sheet_name,i,msg,result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i+1,5).value=msg
        sheet.cell(i+1,7).value=result
        wb.save(file_name)

    @staticmethod
    def write_back(self):
        # wb=load_workbook(project_path.test_path)
        pass
        # sheet_1=wb['init']









if __name__ == '__main__':
    data=Do_excle.get_information(project_path.test_path)
    print(data)
    # for i in data:
    #     print(i['data_2'],type(i['case_id']))
    #     res=requests.get(i['url'],eval(i['data_2']))

        # Do_excle(project_path.test_path,'phone').write_information(i['case_id'],str(res.json()),'pass')
        # res.json()
    # Do_excle.write_information('../test_data/datathing.xlsx','joke',2,'1234','678')